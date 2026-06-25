import io
from datetime import date

from django.contrib.auth.mixins import AccessMixin
from django.contrib import messages
from django.http import HttpResponse


class StaffRequiredMixin(AccessMixin):
    staff_redirect_url = None
    staff_error_message = 'Necesitas ser miembro del staff para acceder a esta página.'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if not request.user.is_staff:
            messages.error(request, self.staff_error_message)
            redirect_url = self.staff_redirect_url or '/shop/'
            from django.shortcuts import redirect
            return redirect(redirect_url)
        return super().dispatch(request, *args, **kwargs)


class ExportMixin:
    """
    Mixin genérico para exportar el queryset filtrado a PDF o Excel.

    Uso en cualquier ListView:
        export_fields = [
            ('Encabezado', 'attr'),           # acceso simple
            ('Relación',  'fk.attr'),          # dotted notation
            ('Calculado', lambda obj: ...),    # callable
        ]
        export_filename = 'nombre_archivo'    # sin extensión
    Activación: añadir ?export=pdf  o  ?export=excel a la URL de listado.
    """

    export_fields: list = []
    export_filename: str = 'listado'

    # ── Interceptor ────────────────────────────────────────────────────
    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('export', '').lower()
        if fmt == 'pdf':
            return self._export_pdf()
        if fmt == 'excel':
            return self._export_excel()
        return super().get(request, *args, **kwargs)

    # ── Helpers internos ───────────────────────────────────────────────
    def get_export_fields(self):
        """
        Devuelve la lista [(label, accessor), ...] a usar en la exportación.
        Las subclases pueden sobreescribir este método para selección dinámica.
        """
        return self.export_fields

    def _get_export_data(self):
        """Devuelve (headers, rows) usando el queryset filtrado (sin paginación)."""
        fields = self.get_export_fields()
        headers = [label for label, _ in fields]
        rows = []
        for obj in self.get_queryset():
            rows.append([self._resolve(obj, acc) for _, acc in fields])
        return headers, rows

    @staticmethod
    def _resolve(obj, accessor):
        """Resuelve un valor de celda: string simple, dotted o callable."""
        if callable(accessor):
            value = accessor(obj)
            return str(value) if value is not None else ''
        value = obj
        for part in accessor.split('.'):
            if value is None:
                return ''
            attr = getattr(value, part, '')
            value = attr() if callable(attr) else attr
        return str(value) if value is not None else ''

    def _filename(self, ext):
        return f'{self.export_filename}_{date.today().strftime("%Y%m%d")}.{ext}'

    # ── PDF ────────────────────────────────────────────────────────────
    def _export_pdf(self):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        )

        headers, rows = self._get_export_data()
        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=2 * cm,
            bottomMargin=1.5 * cm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=14,
            spaceAfter=6,
        )
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#6c757d'),
            spaceAfter=12,
        )

        elements = [
            Paragraph(self.export_filename.replace('_', ' ').title(), title_style),
            Paragraph(
                f'Generado el {date.today().strftime("%d/%m/%Y")} &nbsp;|&nbsp; '
                f'{len(rows)} registro{"s" if len(rows) != 1 else ""}',
                subtitle_style,
            ),
        ]

        # Anchos de columna proporcionales
        page_w = landscape(A4)[0] - 3 * cm
        col_w = page_w / max(len(headers), 1)

        data = [headers] + rows
        table = Table(data, colWidths=[col_w] * len(headers), repeatRows=1)
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#212529')),
            ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, 0),  9),
            # Cuerpo
            ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',      (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            # Alineación y padding
            ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            # Bordes
            ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#dee2e6')),
            ('LINEBELOW',     (0, 0), (-1, 0),  1.5, colors.HexColor('#495057')),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self._filename("pdf")}"'
        return response

    # ── Excel ──────────────────────────────────────────────────────────
    def _export_excel(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        headers, rows = self._get_export_data()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.export_filename.replace('_', ' ').title()[:31]

        # Estilos
        hdr_fill  = PatternFill('solid', fgColor='212529')
        hdr_font  = Font(bold=True, color='FFFFFF', size=10)
        even_fill = PatternFill('solid', fgColor='F8F9FA')
        odd_fill  = PatternFill('solid', fgColor='FFFFFF')
        center    = Alignment(horizontal='center', vertical='center')
        left      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        thin      = Side(style='thin', color='DEE2E6')
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Fila de encabezado
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = center
            cell.border    = border
        ws.row_dimensions[1].height = 20

        # Filas de datos
        for r_idx, row in enumerate(rows, 2):
            fill = even_fill if r_idx % 2 == 0 else odd_fill
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = left
                cell.border    = border
                cell.fill      = fill

        # Ancho de columnas automático (máx 45)
        for c_idx, header in enumerate(headers, 1):
            max_len = len(header)
            for row in rows:
                max_len = max(max_len, len(str(row[c_idx - 1])))
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 4, 45)

        # Fila de totales al pie
        total_row = len(rows) + 2
        ws.cell(row=total_row, column=1, value=f'Total registros: {len(rows)}').font = Font(bold=True, italic=True)

        ws.freeze_panes = 'A2'

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        ct = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response = HttpResponse(buffer, content_type=ct)
        response['Content-Disposition'] = f'attachment; filename="{self._filename("xlsx")}"'
        return response
